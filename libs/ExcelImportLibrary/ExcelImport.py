from builtins import open

import io
import openpyxl
from robot.api.deco import keyword

from .ColumnData import ColumnData
from .RowData import RowData

EMPTY_VALUE: str = ""


class ExcelImport:
    """
    This is a class for reading data in the excel file. We support the excel only XLSX format.
    If you use this class, you would call the method to open the file, select the sheet name first.
    """

    def __init__(self):
        self.__m_file_path = None
        self.__m_workbook = None
        self.__m_working_sheet = None
        self.__m_headers = []
        self.__m_test_cases = []

    @keyword("Open Excel File")
    def open_excel_file(self, file_path: str, data_only: bool = True) -> None:
        """
        Opens the excel file from the relative path in the parameter. This is supported only XLSX file.

        :param data_only:
        :param file_path: The string value that is the relative path file.
               For example c:/program file/python/lib/excel.xlsx
        :raise exception: Cannot open excel file.
        """
        try:

            with open(file_path, "rb") as file_stream:
                excel_file = io.BytesIO(file_stream.read())
                self.__m_workbook = openpyxl.load_workbook(excel_file, keep_vba=True, data_only=data_only)
                self.__m_file_path = file_path
        except FileNotFoundError as exception:
            raise exception

    @keyword("Select Excel Sheet")
    def select_excel_sheet(self, sheet_name: str) -> None:
        """
        The excel must be selected the sheet name before reading the data.

        :param sheet_name: The sheet name in the excel file.
        """
        try:
            self.__clear()
            self.__m_working_sheet = self.__m_workbook[sheet_name]
            self.__set_headers()
        except KeyError as exception:
            raise exception

    def __set_headers(self) -> None:
        max_col = self.get_max_column()
        iter_rows = self.__m_working_sheet.iter_rows(min_row=1, max_col=max_col, max_row=1, values_only=True)
        self.__m_headers = next(iter_rows)

    @keyword("Set Headers Row")
    def set_headers_row(self, headers_row=1) -> None:
        """

        Args:
            headers_row : int (default=1)
        """
        max_col = self.get_max_column()
        iter_rows = self.__m_working_sheet.iter_rows(min_row=headers_row, max_col=max_col, max_row=headers_row,
                                                     values_only=True)
        self.__m_headers = next(iter_rows)

    @keyword("Set Working Rows")
    def set_working_rows(self, start_row: int, end_row: int, max_col: int) -> None:
        """
        Limit the rows that you want to read data on the excel file by fixing the start and end row.
        The working rows maybe contain many test cases. We add the test cases into list. The test case has
        been separated by using first column in that sheet.

        :param start_row: Start row is a first row of the data.
        :param end_row: End row is a last row of the data.
        :param max_col: How many column in that sheet.
        """
        self.__m_test_cases = self.__get_rows_data(
            min_row=start_row,
            max_row=end_row,
            max_col=max_col,
            only_first_data=False
        )

    @keyword("Get Test Case")
    def get_test_case(self, index: int) -> dict:
        """
        Gets the test case in the working row by using the index.

        :param index: The index of the test case.
        :return: A dictionary contains start row, end row, and values. The values are the dictionary that
        key and value are the header name and list of data.
        """
        try:
            return self.__format_test_case(self.__m_test_cases[index])
        except IndexError as exception:
            raise IndexError(str(exception) + ". Please set working rows first.")

    @keyword("Find Test Case")
    def find_test_case(self, sheet_name: str, test_case_name: str, headers_row=1, search_col=1) -> list:
        """
        Find test case by using a name of test case. The test cases maybe more than one test case then
        we add all test cases into list.

        :param sheet_name: string
        :param test_case_name: the name of the test case
        :return: test case list

        Args:
            headers_row: int
        """
        if self.__m_working_sheet is None or self.__m_working_sheet.title != sheet_name:
            self.select_excel_sheet(sheet_name)
        self.set_headers_row(headers_row)
        row_index_list = self.search_row(test_case_name, search_col)
        max_row = self.get_max_row()
        max_col = self.get_max_column()
        test_case_list = []
        for row_index in row_index_list:
            test_case = self.__get_rows_data(
                min_row=row_index,
                max_row=max_row,
                max_col=max_col,
                only_first_data=True
            )
            test_case_list.append(self.__format_test_case(test_case[0]))
        return test_case_list

    @staticmethod
    def __format_test_case(test_case: RowData) -> dict:
        return {
            "start_row": str(test_case.get_start_row()),
            "end_row": test_case.get_end_row(),
            "values": test_case.get_columns_dict()
        }

    @keyword("Get Max Test Cases")
    def get_max_test_cases(self) -> int:
        """
        Gets maximum of test cases in the working rows.

        :return: maximum int number
        """
        return len(self.__m_test_cases)

    @keyword("Search Row")
    def search_row(self, text: str, column: int) -> list:
        """
        Search first row that contain the text.

        :param text: string text
        :param column: the column index that you want to find
        :return: index of the row
        """
        text = self.__clear_text(text)
        row_index = []
        # start two for skipping header
        for row in range(2, self.get_max_row() + 1):
            value = self.__clear_text(self.get_cell(row=row, col=column))
            if value == text:
                row_index.append(row)
        return row_index

    @staticmethod
    def __clear_text(text: str) -> str:
        if text is not None:
            return str(text).lstrip('\"').rstrip('\"').replace(" ", "").replace("\n", "").lower()
        return ""

    def __get_rows_data(self, min_row: int, max_row: int, max_col: int, only_first_data: bool) -> list:
        if self.__m_working_sheet is None:
            raise SyntaxError("Please select the excel sheet before set a working rows.")
        if min_row is None or max_row is None or max_col is None:
            raise ValueError(f'min_row[{min_row}] or max_row[{max_row}] or max_col[{max_col}] should not be None.')
        headers = self.__m_headers
        working_rows = []
        index_row = min_row
        iter_rows = self.__m_working_sheet.iter_rows(
            min_row=index_row,
            max_row=max_row,
            max_col=max_col,
            values_only=True
        )
        for row in iter_rows:
            # set first column to separate the test case
            if row[0] is not None:
                if index_row == 1:
                    headers = self.__get_header()
                else:
                    if only_first_data and len(working_rows) >= 1:
                        break
                    working_rows.append(self.__get_row(row, headers, index_row, max_col))
            else:
                self.__add_next_row(
                    previous_row_data=working_rows[len(working_rows) - 1],
                    row=row,
                    max_col=max_col
                )
            index_row += 1
        return working_rows

    def __get_header(self):
        max_col = self.get_max_column()
        return next(self.__m_working_sheet.iter_rows(
            min_row=1,
            max_col=max_col,
            max_row=1,
            values_only=True)
        )

    @keyword("Get Cell")
    def get_cell(self, row: int, col: int) -> str:
        """
        Gets a data in the row and col.

        :param row: index of the row
        :param col: index of the column
        :return: data in the row and column
        """
        try:
            return self.__m_working_sheet.cell(row, col).value
        except AttributeError as exception:
            self.__throw_attribute_error(exception)

    @keyword("Get Max Row")
    def get_max_row(self) -> int:
        """
        Gets maximum of the row in that sheet.

        :return: int number
        """
        try:
            return self.__m_working_sheet.max_row
        except AttributeError as exception:
            self.__throw_attribute_error(exception)

    @keyword("Get Max Column")
    def get_max_column(self) -> int:
        """
        Gets maximum of the column in that sheet.

        :return: int number
        """
        try:
            return self.__m_working_sheet.max_column
        except AttributeError as exception:
            self.__throw_attribute_error(exception)

    def __throw_attribute_error(self, exception):
        raise AttributeError(str(exception) + ". Please select the excel sheet before set a working rows.")

    def __clear(self) -> None:
        self.__m_working_sheet = None
        self.__m_headers = []
        self.__m_test_cases = []

    @staticmethod
    def __get_row(row, headers, index_row: int, max_col: int):
        row_data = RowData(index_row)
        for index in range(0, max_col):
            column_data = ColumnData()
            if headers:
                column_data.set_header(headers[index])
            column_data.add_value(row[index])
            row_data.add_column(column_data)

        return row_data

    @staticmethod
    def __add_next_row(previous_row_data, row, max_col):
        previous_row_data.increase_row()
        for index in range(0, max_col):
            column_data = previous_row_data.get_columns()[index]
            value = row[index]
            if value is not None:
                column_data.add_value(value)
            else:
                column_data.add_value(EMPTY_VALUE)

    @keyword("To List")
    def to_list(self, text: str, separator=",") -> list:
        text_list = []
        if text is not None:
            for t in text.strip("\n").split(separator):
                text_list.append(t.strip())
        return text_list

    @keyword("Get Sheet Names")
    def get_sheet_names(self) -> list:
        return self.__m_workbook.sheetnames
